"""
AcademicPregrader — Frontend web local
Ejecutar: python app.py
Docker:   docker-compose up --build
"""

import json
import os
import queue
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import unicodedata
import uuid
import webbrowser
from pathlib import Path

try:
    import pty  # POSIX only (Linux/macOS) — no existe en Windows.
except ImportError:
    pty = None

from flask import Flask, Response, jsonify, redirect, render_template, request, session, url_for
from werkzeug.utils import secure_filename
import configparser
import csv as csv_module

from pypdf import PdfWriter

import auth
import code_runner

app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['MAX_CONTENT_LENGTH'] = 512 * 1024 * 1024  # 512 MB upload limit
app.secret_key = auth.get_secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    # Cookie segura en producción (HTTPS en Azure); desactivable para pruebas locales.
    SESSION_COOKIE_SECURE=os.environ.get("PREGRADER_CONFIG_DIR", "").startswith("/app"),
)
auth.init_db()
auth.seed_admin_from_env()

SCRIPT_DIR  = Path(__file__).parent
BACKEND_DIR = SCRIPT_DIR.parent / "backend"
PREGRADER_SCRIPT = BACKEND_DIR / "academic-pregrader.py"

# Config directory: use env var when running in Docker, otherwise backend dir
CONFIG_DIR = Path(os.environ.get("PREGRADER_CONFIG_DIR", str(BACKEND_DIR)))
CONFIG_DIR.mkdir(parents=True, exist_ok=True)

# job_id → Queue de mensajes SSE
_queues: dict[str, queue.Queue] = {}
# job_id → resultados acumulados
_results: dict[str, list] = {}
# job_id → proceso activo
_procs: dict[str, subprocess.Popen] = {}

# run_id → sesión de consola interactiva del preview de código (ver InteractiveRun)
_interactive_runs: dict[str, "InteractiveRun"] = {}


class InteractiveRun:
    """Una ejecución interactiva de código (preview), conectada a un pseudo-terminal
    para que el programa se comporte como si corriera en una consola real: la salida
    se transmite tal cual se produce y el evaluador puede escribir la entrada mientras
    el programa sigue corriendo, sin tener que precargar el stdin de antemano."""

    def __init__(self, run_id: str, language: str, cmd: list[str], tmp_dir: str):
        self.run_id = run_id
        self.language = language
        self.tmp_dir = tmp_dir
        self.queue: queue.Queue = queue.Queue()
        self.done = False
        self.exit_code = None
        self.timed_out = False
        self.started = time.time()
        self._lock = threading.Lock()
        self._finished_at = None

        master_fd, slave_fd = pty.openpty()
        self.master_fd = master_fd
        try:
            self.proc = subprocess.Popen(
                cmd,
                cwd=tmp_dir,
                stdin=slave_fd, stdout=slave_fd, stderr=slave_fd,
                preexec_fn=code_runner._limit_resources if sys.platform != "win32" else None,
                close_fds=True,
                start_new_session=True,
            )
        finally:
            os.close(slave_fd)

        threading.Thread(target=self._pump_output, daemon=True).start()
        threading.Thread(target=self._watchdog, daemon=True).start()

    def _pump_output(self):
        try:
            while True:
                try:
                    chunk = os.read(self.master_fd, 4096)
                except OSError:
                    break  # EIO: el lado esclavo se cerró (el proceso terminó) — fin normal en pty.
                if not chunk:
                    break
                text = chunk.decode("utf-8", errors="replace")
                self.queue.put({"type": "output", "data": text})
        finally:
            self.proc.wait()
            with self._lock:
                self.exit_code = self.proc.returncode
                self.done = True
                self._finished_at = time.time()
            try:
                os.close(self.master_fd)
            except OSError:
                pass
            self.queue.put({
                "type": "done", "exit_code": self.exit_code, "timed_out": self.timed_out,
                "duration_ms": int((time.time() - self.started) * 1000),
            })
            shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def _watchdog(self):
        """Mata el proceso si se pasa del tope de tiempo (evita procesos colgados
        para siempre esperando una entrada que nunca llega, o loops infinitos)."""
        deadline = self.started + code_runner.INTERACTIVE_RUN_TIMEOUT_S
        while time.time() < deadline:
            if self.done:
                return
            time.sleep(1)
        if not self.done:
            self.timed_out = True
            self.kill()

    def write_input(self, text: str) -> bool:
        with self._lock:
            if self.done:
                return False
        try:
            os.write(self.master_fd, (text + "\n").encode("utf-8", errors="replace"))
            return True
        except OSError:
            return False

    def kill(self):
        try:
            os.killpg(self.proc.pid, 9)  # SIGKILL al grupo completo (incluye hijos)
        except Exception:
            try:
                self.proc.kill()
            except Exception:
                pass


def client_ip() -> str:
    """IP real del cliente, respetando el proxy de Azure Container Apps."""
    fwd = request.headers.get("X-Forwarded-For", "")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.remote_addr or "?"


@app.before_request
def _touch_activity():
    """Cualquier request de un usuario ya autenticado cuenta como "actividad real",
    no solo el login: así "último acceso" refleja el uso de la app, no solo cuándo
    se autenticó por última vez (la sesión puede seguir viva por días sin re-login)."""
    auth.touch_last_seen(session.get("user"))


# ==============================
# RUTAS
# ==============================

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        result = auth.authenticate(username, password)
        uname = username.strip().lower()
        if result["status"] == "ok":
            session.clear()
            session["user"] = uname
            auth.record_login(uname)
            auth.log_event(uname, client_ip(), "login")
            if result["user"]["must_change_password"]:
                return redirect(url_for("change_password"))
            next_url = request.form.get("next", "") or url_for("index")
            # Evita open-redirects: solo rutas internas.
            if not next_url.startswith("/"):
                next_url = url_for("index")
            return redirect(next_url)
        if result["status"] == "blocked":
            auth.log_event(uname, client_ip(), "login_blocked")
            error = "Tu usuario está bloqueado. Contacta al administrador."
        else:
            auth.log_event(uname or None, client_ip(), "login_failed", detail=uname)
            error = "Usuario o contraseña incorrectos."
        return render_template("login.html", error=error,
                               next_url=request.form.get("next", "")), 401
    return render_template("login.html", error=None, next_url=request.args.get("next", ""))


@app.route("/logout", methods=["POST", "GET"])
def logout():
    user = session.get("user")
    if user:
        auth.log_event(user, client_ip(), "logout")
    session.clear()
    return redirect(url_for("login"))


@app.route("/change-password", methods=["GET", "POST"], endpoint="change_password")
@auth.login_required
def change_password():
    user = session.get("user")
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        must_change = bool(auth.get_user(user)["must_change_password"])
        # En cambio forzado (primer ingreso) no exigimos la clave actual de nuevo.
        if not must_change and not auth.authenticate(user, current)["status"] == "ok":
            return render_template("change_password.html", error="La contraseña actual es incorrecta.",
                                   must_change=must_change), 400
        if len(new) < 8:
            return render_template("change_password.html", error="La nueva contraseña debe tener al menos 8 caracteres.",
                                   must_change=must_change), 400
        if new != confirm:
            return render_template("change_password.html", error="Las contraseñas no coinciden.",
                                   must_change=must_change), 400
        auth.set_password(user, new)
        auth.log_event(user, client_ip(), "password_changed")
        return redirect(url_for("index"))
    must_change = bool(auth.get_user(user)["must_change_password"])
    return render_template("change_password.html", error=None, must_change=must_change)


@app.route("/")
@auth.login_required
def index():
    username = session["user"]
    user = auth.get_user(username)
    is_admin = bool(user["is_admin"])
    active_course = auth.get_user_active_course(username)
    # Los admins eligen libremente cualquier curso; el resto usa el asignado.
    courses = auth.list_courses() if is_admin else []
    return render_template(
        "index.html",
        is_admin=is_admin,
        active_course=active_course["name"] if active_course else "",
        active_course_id=active_course["id"] if active_course else "",
        courses=courses,
    )


@app.route("/run", methods=["POST"])
@auth.login_required
def run_pregrader():
    """Acepta ZIP + uno o varios PDF como multipart/form-data y lanza la evaluación."""
    zip_file = request.files.get("zip_file")
    pdf_files = request.files.getlist("pdf_file") or request.files.getlist("pdf_files")
    pdf_files = [f for f in pdf_files if f and f.filename]

    if not zip_file or not zip_file.filename:
        return jsonify({"error": "Debes seleccionar el archivo ZIP de entregas."}), 400
    if not pdf_files:
        return jsonify({"error": "El enunciado (PDF) es obligatorio."}), 400

    # Guardar los archivos subidos en un directorio temporal
    tmp_dir = tempfile.mkdtemp()
    zip_name = secure_filename(zip_file.filename) or "entregas.zip"
    zip_path = os.path.join(tmp_dir, zip_name)
    zip_file.save(zip_path)

    pdf_paths = []
    for index, pdf_file in enumerate(pdf_files, start=1):
        original_name = secure_filename(pdf_file.filename) or f"enunciado_{index}.pdf"
        safe_name = f"{index:02d}_{original_name}"
        pdf_path = os.path.join(tmp_dir, safe_name)
        pdf_file.save(pdf_path)
        pdf_paths.append(pdf_path)

    enunciado_path = pdf_paths[0]
    if len(pdf_paths) > 1:
        merged_pdf = os.path.join(tmp_dir, "enunciado-combinado.pdf")
        writer = PdfWriter()
        try:
            for pdf_path in pdf_paths:
                writer.append(pdf_path)
            with open(merged_pdf, "wb") as fh:
                writer.write(fh)
        finally:
            writer.close()
        enunciado_path = merged_pdf

    extra_notes = (request.form.get("extra_notes") or "").strip()[:4000]

    username = session["user"]
    auth.log_event(username, client_ip(), "grade_run", detail=zip_name)

    # Config efectiva del job: global (admin) + personal del usuario.
    job_cfg_dir = build_job_config_dir(username)

    job_id = os.urandom(8).hex()
    q: queue.Queue = queue.Queue()
    _queues[job_id] = q

    def worker():
        cmd = [sys.executable, str(PREGRADER_SCRIPT), zip_path, enunciado_path]
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"
        env["PREGRADER_CONFIG_DIR"] = job_cfg_dir
        if extra_notes:
            env["PREGRADER_EXTRA_NOTES"] = extra_notes

        job_results = []

        try:
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
                cwd=str(BACKEND_DIR),
                env=env,
            )
            _procs[job_id] = proc
            for line in iter(proc.stdout.readline, ""):
                stripped = line.rstrip()
                if stripped.startswith("##STUDENTS## "):
                    try:
                        students = json.loads(stripped[len("##STUDENTS## "):])
                        q.put({"type": "students_list", "students": students})
                    except Exception:
                        q.put({"type": "log", "data": stripped})
                elif stripped.startswith("##RESULT## "):
                    try:
                        result = json.loads(stripped[len("##RESULT## "):])
                        job_results.append(result)
                        q.put({"type": "student_result", "data": result})
                    except Exception:
                        q.put({"type": "log", "data": stripped})
                elif stripped.startswith("##DONE## "):
                    pass
                else:
                    q.put({"type": "log", "data": stripped})
            proc.stdout.close()
            proc.wait()
        except Exception as exc:
            q.put({"type": "log", "data": f"[ERROR interno] {exc}"})
        finally:
            _procs.pop(job_id, None)
            # Limpiar archivos temporales subidos y la config del job
            import shutil
            shutil.rmtree(tmp_dir, ignore_errors=True)
            shutil.rmtree(job_cfg_dir, ignore_errors=True)

        cancelled = getattr(proc, "_cancelled", False) if "proc" in dir() else False
        if cancelled:
            q.put({"type": "cancelled"})
            q.put(None)
            return

        _results[job_id] = job_results
        if job_results:
            try:
                auth.save_grading_history(username, job_id, zip_name, job_results)
            except Exception as exc:
                print(f"[historial] No se pudo guardar el historial de {username}: {exc}")
        q.put({"type": "done", "job_id": job_id, "count": len(job_results)})
        q.put(None)

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/stop/<job_id>", methods=["POST"])
@auth.login_required
def stop_job(job_id: str):
    proc = _procs.get(job_id)
    if proc is None:
        return jsonify({"ok": False, "error": "Job no encontrado o ya finalizado."}), 404
    try:
        proc._cancelled = True
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    q = _queues.get(job_id)
    if q:
        q.put({"type": "cancelled"})
        q.put(None)
    return jsonify({"ok": True})


@app.route("/stream/<job_id>")
@auth.login_required
def stream(job_id: str):
    def generate():
        q = _queues.get(job_id)
        if q is None:
            yield f"data: {json.dumps({'type': 'error', 'data': 'Job no encontrado.'})}\n\n"
            return
        while True:
            try:
                msg = q.get(timeout=120)
            except queue.Empty:
                yield 'data: {"type":"keepalive"}\n\n'
                continue
            if msg is None:
                _queues.pop(job_id, None)
                break
            yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.route("/download/<job_id>")
@auth.login_required
def download(job_id: str):
    results = _results.get(job_id)
    if not results:
        return "No hay resultados para este job.", 404

    auth.log_event(session.get("user"), client_ip(), "download_csv", detail=job_id)

    def clean(v):
        return str(v).replace("\r\n", " ").replace("\r", " ").replace("\n", " ").strip()

    import io
    output = io.StringIO()
    writer = csv_module.writer(output, quoting=csv_module.QUOTE_ALL)
    writer.writerow(["Estudiante", "Compila", "Plagio", "Nota", "Comentario"])
    for r in results:
        writer.writerow([
            clean(r.get("estudiante", "")),
            clean(r.get("compila",    "")),
            clean(r.get("plagio",     "")),
            clean(r.get("nota",       "")),
            clean(r.get("comentario", "")),
        ])

    return Response(
        "﻿" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=resultados.csv"},
    )


@app.route("/download/<job_id>/excel")
@auth.login_required
def download_excel(job_id: str):
    results = _results.get(job_id)
    if not results:
        return "No hay resultados para este job.", 404

    auth.log_event(session.get("user"), client_ip(), "download_excel", detail=job_id)

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    headers = ["Estudiante", "Compila", "Plagio", "Nota", "Comentario"]

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1E293B")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="334155")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border

    green_fill  = PatternFill("solid", fgColor="D1FAE5")
    red_fill    = PatternFill("solid", fgColor="FEE2E2")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")
    orange_fill = PatternFill("solid", fgColor="FFEDD5")

    for r in results:
        compila  = str(r.get("compila", ""))
        plagio   = str(r.get("plagio",  ""))
        nota_raw = r.get("nota", "")
        try:
            nota_val = float(nota_raw)
        except (ValueError, TypeError):
            nota_val = None

        row = [
            str(r.get("estudiante", "")),
            compila,
            plagio,
            nota_val if nota_val is not None else str(nota_raw),
            str(r.get("comentario", "")),
        ]
        ws.append(row)
        row_idx = ws.max_row

        c_compila = ws.cell(row=row_idx, column=2)
        c_compila.alignment = Alignment(wrap_text=True, vertical="top")
        if compila == "SI":
            c_compila.fill = green_fill
        elif compila and compila != "N/A":
            c_compila.fill = red_fill

        c_plagio = ws.cell(row=row_idx, column=3)
        if plagio:
            c_plagio.fill = orange_fill

        c_nota = ws.cell(row=row_idx, column=4)
        c_nota.alignment    = Alignment(horizontal="center", vertical="center")
        c_nota.number_format = "0.0"
        if nota_val is not None:
            if nota_val >= 3.5:
                c_nota.fill = green_fill
            elif nota_val >= 2.5:
                c_nota.fill = yellow_fill
            else:
                c_nota.fill = red_fill

        ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = border

    col_widths = [32, 48, 30, 8, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=resultados.xlsx"},
    )


def _name_tokens(name: str) -> set:
    """Tokens normalizados de un nombre (sin acentos, mayúsculas, sin iniciales)."""
    s = unicodedata.normalize("NFKD", str(name))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Za-z0-9 ]", " ", s).upper()
    return {t for t in s.split() if len(t) > 1}


def _match_roster(student_name: str, roster_tokens: list) -> dict | None:
    """Empareja el nombre calificado con una fila del roster por tokens de nombre."""
    g = _name_tokens(student_name)
    if not g:
        return None
    best, best_score = None, 0.0
    for entry, tokens in roster_tokens:
        if not tokens:
            continue
        inter = g & tokens
        if not inter:
            continue
        subset = g <= tokens or tokens <= g
        score = 1.0 if subset else len(inter) / len(g | tokens)
        if score > best_score:
            best_score, best = score, entry
    return best if best_score >= 0.5 else None


@app.route("/download/<job_id>/campus")
@auth.login_required
def download_campus(job_id: str):
    """Exporta las notas en el formato de importación del Campus Virtual (D2L)."""
    results = _results.get(job_id)
    if not results:
        return "No hay resultados para este job.", 404

    if not _read_global_cfg().getboolean("features", "campus_export_enabled", fallback=False):
        return jsonify({"error": "La exportación al Campus Virtual está deshabilitada por el administrador."}), 403

    course = auth.get_user_active_course(session["user"])
    if not course:
        return jsonify({"error": "No tienes un curso activo. Selecciona o pide que te asignen uno."}), 400

    roster = auth.get_course_students(course["id"])
    if not roster:
        return jsonify({"error": "El curso activo no tiene estudiantes cargados."}), 400

    item = (request.args.get("item", "") or "").strip() or "Nota"
    roster_tokens = [(s, _name_tokens(s.get("full_name", ""))) for s in roster]

    matched_rows, unmatched = [], []
    for r in results:
        name = str(r.get("estudiante", ""))
        nota = r.get("nota", "")
        try:
            nota_val = round(float(nota), 1)
        except (ValueError, TypeError):
            nota_val = None
        entry = _match_roster(name, roster_tokens)
        org_id = (entry or {}).get("org_id", "").strip() if entry else ""
        if entry and org_id and nota_val is not None:
            matched_rows.append((org_id, nota_val))
        else:
            unmatched.append(name)

    import io
    output = io.StringIO()
    writer = csv_module.writer(output)
    writer.writerow(["OrgDefinedId", f"{item} Points Grade", "End-of-Line Indicator"])
    for org_id, nota_val in matched_rows:
        writer.writerow([org_id, f"{nota_val:.1f}", "#"])

    auth.log_event(session.get("user"), client_ip(), "download_campus",
                   detail=f"{course['name']} · {item} · {len(matched_rows)} ok / {len(unmatched)} sin emparejar")

    from urllib.parse import quote as _urlquote
    safe_item = re.sub(r"[^A-Za-z0-9_-]+", "_", item).strip("_") or "notas"
    resp = Response(
        "\ufeff" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=campus_{safe_item}.csv"},
    )
    resp.headers["X-Matched-Count"] = str(len(matched_rows))
    resp.headers["X-Unmatched-Count"] = str(len(unmatched))
    resp.headers["X-Unmatched"] = _urlquote(" · ".join(unmatched[:40]))
    return resp


@app.route("/results/<job_id>", methods=["POST"])
@auth.login_required
def update_results(job_id: str):
    """Persiste las notas/comentarios editados en la interfaz para las descargas."""
    if job_id not in _results:
        return jsonify({"ok": False, "error": "Job no encontrado."}), 404
    data = request.get_json(silent=True)
    if not isinstance(data, list):
        return jsonify({"ok": False, "error": "Formato inválido."}), 400
    # Preserva el código fuente ya calculado (el cliente no lo reenvía en cada sync).
    previous_by_name = {r.get("estudiante", ""): r for r in _results.get(job_id, [])}
    cleaned = []
    for item in data[:2000]:
        if not isinstance(item, dict):
            continue
        estudiante = str(item.get("estudiante", ""))[:300]
        cleaned.append({
            "estudiante": estudiante,
            "compila":    str(item.get("compila", ""))[:4000],
            "plagio":     str(item.get("plagio", ""))[:300],
            "nota":       item.get("nota", ""),
            "comentario": str(item.get("comentario", ""))[:5000],
            "codigo":     previous_by_name.get(estudiante, {}).get("codigo", ""),
        })
    _results[job_id] = cleaned
    try:
        auth.update_grading_history_results(session["user"], job_id, cleaned)
    except Exception as exc:
        print(f"[historial] No se pudo sincronizar el historial: {exc}")
    return jsonify({"ok": True, "count": len(cleaned)})


@app.route("/history", methods=["GET"])
@auth.login_required
def get_history():
    """Lista (metadatos) de las últimas calificaciones guardadas del usuario actual."""
    items = auth.get_grading_history(session["user"])
    return jsonify({"items": items})


@app.route("/history/<job_id>", methods=["GET"])
@auth.login_required
def get_history_item(job_id: str):
    """Snapshot completo de una calificación guardada, para retomarla tal cual quedó."""
    item = auth.get_grading_history_item(session["user"], job_id)
    if not item:
        return jsonify({"error": "No se encontró esa calificación en tu historial."}), 404
    # Repuebla la caché en memoria para que descargas (CSV/Excel/Campus) sigan funcionando.
    _results[job_id] = item["results"]
    return jsonify(item)


# ==============================
# CONFIGURACIÓN
# ==============================

def _cfg_path() -> Path:
    return CONFIG_DIR / "config.ini"


# Valores por defecto de la config personal (por usuario).
USER_CONFIG_DEFAULTS = {
    "submission_language": "auto",
    "enable_compilation":  True,
    "enable_plagiarism":   False,
    "enable_llm":          True,
    "enable_cache":        False,
    "comment_on_max":      False,
    "llm_provider":        "openai",
    "llm_model":           "gpt-5-mini",
    "threshold":           0.7,
}


def _read_global_cfg() -> configparser.RawConfigParser:
    cfg = configparser.RawConfigParser()
    if _cfg_path().exists():
        cfg.read(_cfg_path(), encoding="utf-8")
    return cfg


def _has_api_key(cfg: configparser.RawConfigParser, provider: str) -> bool:
    env_name = f"{provider.upper()}_API_KEY"
    config_name = f"{provider.lower()}_api_key"
    return bool(
        os.environ.get(env_name, "").strip()
        or cfg.get("llm", config_name, fallback="").strip()
    )


def _provider_is_available(cfg: configparser.RawConfigParser, provider: str) -> bool:
    return (
        provider in {"openai", "gemini"}
        and cfg.getboolean("llm", f"{provider}_enabled", fallback=True)
        and _has_api_key(cfg, provider)
    )


def _available_providers(cfg: configparser.RawConfigParser) -> list[str]:
    return [
        provider
        for provider in ("openai", "gemini")
        if _provider_is_available(cfg, provider)
    ]


def _provider_defaults(provider: str) -> tuple[str, str]:
    if provider == "gemini":
        return "gemini", "gemini-2.0-flash"
    return "openai", "gpt-5-mini"


def _effective_provider(cfg: configparser.RawConfigParser, requested: str) -> tuple[str, str | None]:
    available = _available_providers(cfg)
    if requested in available:
        return requested, None
    if available:
        provider = available[0]
        return provider, _provider_defaults(provider)[1]
    return "openai", "gpt-5-mini"


def _default_jplag() -> str:
    for jar in sorted(BACKEND_DIR.glob("jplag*.jar")):
        return str(jar)
    return ""


def _effective_user_config(username: str) -> dict:
    """Config personal del usuario mezclada con los valores por defecto."""
    data = {**USER_CONFIG_DEFAULTS, **auth.get_user_config(username)}
    return data


@app.route("/config", methods=["GET"])
@auth.login_required
def get_config():
    """Config personal del usuario. Nunca expone las claves API (solo si existen)."""
    cfg = _read_global_cfg()
    uc = _effective_user_config(session["user"])
    provider, fallback_model = _effective_provider(cfg, str(uc["llm_provider"]))
    return jsonify({
        **uc,
        "llm_provider": provider,
        "llm_model": fallback_model or uc["llm_model"],
        # Estado de las claves (gestionadas por el admin), sin revelar su valor.
        "gemini_has_api_key": _has_api_key(cfg, "gemini"),
        "openai_has_api_key": _has_api_key(cfg, "openai"),
        "gemini_available": _provider_is_available(cfg, "gemini"),
        "openai_available": _provider_is_available(cfg, "openai"),
        "campus_export_enabled": cfg.getboolean("features", "campus_export_enabled", fallback=False),
    })


@app.route("/config", methods=["POST"])
@auth.login_required
def save_config():
    """Guarda solo la config personal; ignora cualquier campo de administrador."""
    data = request.get_json(force=True) or {}
    cfg = _read_global_cfg()
    requested_provider = str(data.get("llm_provider", "openai"))
    provider, fallback_model = _effective_provider(cfg, requested_provider)
    saved = {
        "submission_language": str(data.get("submission_language", "auto")),
        "enable_compilation":  bool(data.get("enable_compilation", True)),
        "enable_plagiarism":   bool(data.get("enable_plagiarism", False)),
        "enable_llm":          bool(data.get("enable_llm", True)) and bool(_available_providers(cfg)),
        "enable_cache":        bool(data.get("enable_cache", False)),
        "comment_on_max":      bool(data.get("comment_on_max", False)),
        "llm_provider":        provider,
        "llm_model":           fallback_model or str(data.get("llm_model", _provider_defaults(provider)[1])),
        "threshold":           float(data.get("threshold", 0.7) or 0.7),
    }
    auth.save_user_config(session["user"], saved)
    return jsonify({"ok": True})


def _is_openai_reasoning_model(model: str) -> bool:
    """Modelos GPT-5/o-series: no aceptan temperature/seed personalizados y usan
    max_completion_tokens en vez de max_tokens (Chat Completions API)."""
    m = (model or "").strip().lower()
    return m.startswith(("gpt-5", "o1", "o3", "o4"))


def _call_llm_question(provider: str, model: str, api_key: str, prompt: str) -> str | None:
    if provider == "gemini":
        try:
            import google.generativeai as genai
            genai.configure(api_key=api_key)
            m = genai.GenerativeModel(model_name=model, generation_config={"temperature": 0.3, "max_output_tokens": 1024})
            return m.generate_content(prompt).text
        except Exception:
            return None
    elif provider == "openai":
        try:
            from openai import OpenAI
            call_kwargs = {"model": model, "messages": [{"role": "user", "content": prompt}]}
            if _is_openai_reasoning_model(model):
                call_kwargs["max_completion_tokens"] = 1024
                call_kwargs["reasoning_effort"] = "low"
            else:
                call_kwargs["temperature"] = 0.3
                call_kwargs["max_tokens"] = 1024
            r = OpenAI(api_key=api_key).chat.completions.create(**call_kwargs)
            return r.choices[0].message.content
        except Exception:
            return None
    return None


@app.route("/ask_ai", methods=["POST"])
@auth.login_required
def ask_ai():
    data = request.get_json(force=True) or {}
    code     = str(data.get("code",     "")).strip()[:80000]
    question = str(data.get("question", "")).strip()[:2000]
    if not question:
        return jsonify({"error": "Pregunta vacía."}), 400

    cfg = _read_global_cfg()
    uc  = _effective_user_config(session["user"])
    provider, fallback_model = _effective_provider(cfg, str(uc["llm_provider"]))
    model = fallback_model or str(uc["llm_model"])

    if provider == "gemini":
        api_key = os.environ.get("GEMINI_API_KEY", "").strip() or cfg.get("llm", "gemini_api_key", fallback="")
    else:
        api_key = os.environ.get("OPENAI_API_KEY", "").strip() or cfg.get("llm", "openai_api_key", fallback="")

    if not api_key:
        return jsonify({"error": "No hay API key configurada."}), 503

    code_section = f"\nCÓDIGO DEL ESTUDIANTE:\n{code}\n" if code else ""
    prompt = ("Eres un asistente para un evaluador universitario que revisa entregas de código.\n"
              "Responde de forma concisa y directa. No des una calificación — solo responde la pregunta."
              + code_section +
              f"\nPREGUNTA DEL EVALUADOR:\n{question}")

    answer = _call_llm_question(provider, model, api_key, prompt)
    if answer is None:
        return jsonify({"error": "El LLM no pudo responder. Intenta de nuevo."}), 503
    return jsonify({"answer": answer})


@app.route("/run_code", methods=["POST"])
@auth.login_required
def run_code_endpoint():
    """Compila y ejecuta el código de una entrega (Python/C++/Java) y devuelve la consola.
    Modo "todo de una" (no interactivo); se mantiene por compatibilidad. El preview de
    código en el frontend usa /run_code/start (consola interactiva de verdad)."""
    data = request.get_json(force=True) or {}
    code = str(data.get("code", ""))[:200_000]
    stdin_text = str(data.get("stdin", ""))[:20_000]
    if not code.strip():
        return jsonify({"ok": False, "error": "No hay código para ejecutar."}), 400

    auth.log_event(session.get("user"), client_ip(), "run_code")
    try:
        result = code_runner.run_code(code, stdin_text)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error interno al ejecutar el código: {e}"}), 500
    return jsonify(result)


@app.route("/run_code/start", methods=["POST"])
@auth.login_required
def run_code_start():
    """Compila (si aplica) y arranca una consola interactiva real: la salida se
    transmite según se produce y se puede enviar entrada mientras el programa corre,
    en vez de precargar el stdin antes de ejecutar."""
    if pty is None:
        return jsonify({"ok": False, "error": "La consola interactiva no está disponible en este servidor (requiere Linux/macOS)."}), 501

    data = request.get_json(force=True) or {}
    code = str(data.get("code", ""))[:200_000]
    if not code.strip():
        return jsonify({"ok": False, "error": "No hay código para ejecutar."}), 400

    auth.log_event(session.get("user"), client_ip(), "run_code")
    try:
        prep = code_runner.prepare_interactive(code)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error interno al preparar la ejecución: {e}"}), 500

    if not prep.get("ok"):
        shutil.rmtree(prep.pop("tmp_dir", ""), ignore_errors=True)
        return jsonify(prep)

    run_id = uuid.uuid4().hex
    try:
        run = InteractiveRun(run_id, prep["language"], prep["cmd"], prep["tmp_dir"])
    except Exception as e:
        shutil.rmtree(prep["tmp_dir"], ignore_errors=True)
        return jsonify({"ok": False, "error": f"No se pudo iniciar el proceso: {e}"}), 500
    _interactive_runs[run_id] = run
    return jsonify({"ok": True, "run_id": run_id, "language": prep["language"]})


@app.route("/run_code/stream/<run_id>")
@auth.login_required
def run_code_stream(run_id: str):
    def generate():
        run = _interactive_runs.get(run_id)
        if run is None:
            yield f"data: {json.dumps({'type': 'error', 'data': 'Sesión no encontrada.'})}\n\n"
            return
        while True:
            try:
                msg = run.queue.get(timeout=60)
            except queue.Empty:
                yield 'data: {"type":"keepalive"}\n\n'
                continue
            yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
            if msg.get("type") == "done":
                _interactive_runs.pop(run_id, None)
                break

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/run_code/input/<run_id>", methods=["POST"])
@auth.login_required
def run_code_input(run_id: str):
    run = _interactive_runs.get(run_id)
    if run is None:
        return jsonify({"ok": False, "error": "Sesión no encontrada o ya terminada."}), 404
    data = request.get_json(force=True) or {}
    text = str(data.get("text", ""))[:4000]
    if not run.write_input(text):
        return jsonify({"ok": False, "error": "El programa ya terminó."}), 409
    return jsonify({"ok": True})


@app.route("/run_code/stop/<run_id>", methods=["POST"])
@auth.login_required
def run_code_stop(run_id: str):
    run = _interactive_runs.get(run_id)
    if run is None:
        return jsonify({"ok": False, "error": "Sesión no encontrada o ya terminada."}), 404
    run.kill()
    return jsonify({"ok": True})


def build_job_config_dir(username: str) -> str:
    """Crea un dir temporal con un config.ini que fusiona la config global (admin)
    y la personal del usuario, para pasarlo al backend sin pisar la config global."""
    g = _read_global_cfg()
    uc = _effective_user_config(username)
    provider, fallback_model = _effective_provider(g, str(uc["llm_provider"]))
    llm_available = bool(_available_providers(g))
    cfg = configparser.RawConfigParser()

    cfg.add_section("steps")
    cfg.set("steps", "enable_compilation", "true" if uc["enable_compilation"] else "false")
    cfg.set("steps", "enable_plagiarism",  "true" if uc["enable_plagiarism"]  else "false")
    cfg.set("steps", "enable_llm",         "true" if uc["enable_llm"] and llm_available else "false")
    cfg.set("steps", "enable_cache",       "true" if uc["enable_cache"]       else "false")

    cfg.add_section("llm")
    cfg.set("llm", "provider", provider)                                  # usuario
    cfg.set("llm", "model",    fallback_model or str(uc["llm_model"]))    # usuario
    cfg.set("llm", "gemini_api_key", g.get("llm", "gemini_api_key", fallback=""))   # admin
    cfg.set("llm", "openai_api_key", g.get("llm", "openai_api_key", fallback=""))   # admin

    cfg.add_section("paths")
    cfg.set("paths", "jplag_jar", g.get("paths", "jplag_jar", fallback=_default_jplag()))  # admin

    cfg.add_section("plagiarism")
    cfg.set("plagiarism", "threshold", str(uc["threshold"]))             # usuario

    cfg.add_section("performance")
    cfg.set("performance", "max_workers", g.get("performance", "max_workers", fallback="3"))  # admin

    cfg.add_section("prompt")
    cfg.set("prompt", "system_instruction", g.get("prompt", "system_instruction", fallback=""))  # admin
    cfg.set("prompt", "eval_rules",         g.get("prompt", "eval_rules",         fallback=""))   # admin

    cfg.add_section("submission")
    cfg.set("submission", "language", uc["submission_language"])         # usuario
    cfg.set("submission", "comment_on_max", "true" if uc["comment_on_max"] else "false")  # usuario

    job_dir = tempfile.mkdtemp(prefix="pgcfg_")
    with open(os.path.join(job_dir, "config.ini"), "w", encoding="utf-8") as f:
        cfg.write(f)
    return job_dir


# ==============================
# ADMIN
# ==============================

@app.route("/admin")
@auth.admin_required
def admin_dashboard():
    return render_template("admin.html")


@app.route("/admin/api/config", methods=["GET"])
@auth.admin_required
def admin_get_config():
    """Config global (solo admin). No devuelve el valor de las claves, solo su estado."""
    cfg = _read_global_cfg()
    return jsonify({
        "gemini_has_api_key": _has_api_key(cfg, "gemini"),
        "openai_has_api_key": _has_api_key(cfg, "openai"),
        "gemini_enabled":     cfg.getboolean("llm", "gemini_enabled", fallback=True),
        "openai_enabled":     cfg.getboolean("llm", "openai_enabled", fallback=True),
        "jplag_jar":          cfg.get("paths",       "jplag_jar",          fallback=_default_jplag()),
        "max_workers":        cfg.getint("performance", "max_workers",     fallback=4),
        "campus_export_enabled": cfg.getboolean("features", "campus_export_enabled", fallback=False),
        "system_instruction": cfg.get("prompt",      "system_instruction", fallback=""),
        "eval_rules":         cfg.get("prompt",      "eval_rules",         fallback=""),
    })


@app.route("/admin/api/config", methods=["POST"])
@auth.admin_required
def admin_save_config():
    data = request.get_json(force=True) or {}
    cfg = _read_global_cfg()

    def ensure(section):
        if not cfg.has_section(section):
            cfg.add_section(section)

    ensure("llm")
    cfg.set("llm", "gemini_enabled", "true" if bool(data.get("gemini_enabled", True)) else "false")
    cfg.set("llm", "openai_enabled", "true" if bool(data.get("openai_enabled", True)) else "false")
    new_gemini_key = str(data.get("gemini_api_key", "")).strip()
    new_openai_key = str(data.get("openai_api_key", "")).strip()
    if new_gemini_key and not os.environ.get("GEMINI_API_KEY", "").strip():
        cfg.set("llm", "gemini_api_key", new_gemini_key)
    if new_openai_key and not os.environ.get("OPENAI_API_KEY", "").strip():
        cfg.set("llm", "openai_api_key", new_openai_key)
    ensure("paths")
    cfg.set("paths", "jplag_jar", str(data.get("jplag_jar", "")))
    ensure("performance")
    cfg.set("performance", "max_workers", str(int(data.get("max_workers", 3))))
    ensure("features")
    cfg.set("features", "campus_export_enabled", "true" if bool(data.get("campus_export_enabled", False)) else "false")
    ensure("prompt")
    cfg.set("prompt", "system_instruction", str(data.get("system_instruction", "")).strip())
    cfg.set("prompt", "eval_rules",         str(data.get("eval_rules",         "")).strip())

    with open(_cfg_path(), "w", encoding="utf-8") as f:
        cfg.write(f)
    auth.log_event(session.get("user"), client_ip(), "admin_save_config")
    return jsonify({"ok": True})


@app.route("/admin/api/overview")
@auth.admin_required
def admin_overview():
    return jsonify({
        "users": auth.get_users_overview(),
        "events": auth.get_events(limit=500),
        "courses": auth.list_courses(),
    })


@app.route("/admin/api/events/clear", methods=["POST"])
@auth.admin_required
def admin_clear_events():
    """Borra los eventos de calificaciones y descargas (reinicia esos contadores)."""
    types = ["grade_run", "download_csv", "download_excel", "download_campus"]
    deleted = auth.delete_events(types)
    auth.log_event(session.get("user"), client_ip(), "admin_clear_events", detail=str(deleted))
    return jsonify({"ok": True, "deleted": deleted})


@app.route("/admin/api/users", methods=["POST"])
@auth.admin_required
def admin_create_user():
    data = request.get_json(force=True) or {}
    username = str(data.get("username", "")).strip().lower()
    is_admin = bool(data.get("is_admin", False))
    if not username:
        return jsonify({"error": "El nombre de usuario es obligatorio."}), 400
    if auth.get_user(username):
        return jsonify({"error": "Ese usuario ya existe."}), 409
    # Clave temporal: el usuario la cambia en su primer ingreso.
    temp = auth.generate_temp_password()
    auth.create_user(username, temp, is_admin=is_admin, must_change=True)
    auth.log_event(session.get("user"), client_ip(), "admin_create_user", detail=username)
    return jsonify({"ok": True, "username": username, "temp_password": temp})


@app.route("/admin/api/users/<username>/block", methods=["POST"])
@auth.admin_required
def admin_block(username: str):
    data = request.get_json(force=True) or {}
    blocked = bool(data.get("blocked", True))
    if username.strip().lower() == session.get("user") and blocked:
        return jsonify({"error": "No puedes bloquearte a ti mismo."}), 400
    if not auth.set_blocked(username, blocked):
        return jsonify({"error": "Usuario no encontrado."}), 404
    auth.log_event(session.get("user"), client_ip(),
                   "admin_block" if blocked else "admin_unblock", detail=username.strip().lower())
    return jsonify({"ok": True})


@app.route("/admin/api/users/<username>/reset-password", methods=["POST"])
@auth.admin_required
def admin_reset_password(username: str):
    temp = auth.reset_password(username)
    if temp is None:
        return jsonify({"error": "Usuario no encontrado."}), 404
    auth.log_event(session.get("user"), client_ip(), "admin_reset_password", detail=username.strip().lower())
    return jsonify({"ok": True, "temp_password": temp})


@app.route("/admin/api/users/<username>/admin", methods=["POST"])
@auth.admin_required
def admin_set_role(username: str):
    data = request.get_json(force=True) or {}
    make_admin = bool(data.get("is_admin", False))
    uname = username.strip().lower()
    # No permitir quedarse sin admins.
    if not make_admin and auth.get_user(uname) and auth.get_user(uname)["is_admin"] and auth.count_admins() <= 1:
        return jsonify({"error": "Debe existir al menos un administrador."}), 400
    if not auth.set_admin(uname, make_admin):
        return jsonify({"error": "Usuario no encontrado."}), 404
    auth.log_event(session.get("user"), client_ip(), "admin_set_role", detail=f"{uname}:{make_admin}")
    return jsonify({"ok": True})


@app.route("/admin/api/users/<username>", methods=["DELETE"])
@auth.admin_required
def admin_delete_user(username: str):
    uname = username.strip().lower()
    if uname == session.get("user"):
        return jsonify({"error": "No puedes eliminar tu propio usuario."}), 400
    if auth.get_user(uname) and auth.get_user(uname)["is_admin"] and auth.count_admins() <= 1:
        return jsonify({"error": "Debe existir al menos un administrador."}), 400
    if not auth.delete_user(uname):
        return jsonify({"error": "Usuario no encontrado."}), 404
    auth.log_event(session.get("user"), client_ip(), "admin_delete_user", detail=uname)
    return jsonify({"ok": True})


# ==============================
# CURSOS / CLASES
# ==============================

_ORG_ID_RE = re.compile(r"^\d{6,15}$")


def parse_class_list(text: str) -> list[dict]:
    """Convierte la lista de clase pegada del Campus Virtual en estudiantes.

    Formato esperado (columnas copiadas de la tabla del Campus): por cada
    estudiante -> Nombre (APELLIDOS, Nombres), Username, OrgDefinedId (id
    numérico), Rol y Fecha, separados por tabuladores y filas por saltos de
    línea. Al copiar desde el navegador las celdas quedan separadas por TAB;
    aquí normalizamos TAB a salto de línea y anclamos cada registro en el id
    numérico, tomando el username y el nombre inmediatamente anteriores.
    """
    if not text:
        return []
    # Cada celda copiada de una tabla queda separada por TAB; la volvemos línea.
    raw = text.replace("\t", "\n")
    lines = [ln.strip() for ln in raw.splitlines()]
    lines = [ln for ln in lines if ln]

    students: list[dict] = []
    seen: set[str] = set()
    for i, line in enumerate(lines):
        if not _ORG_ID_RE.match(line):
            continue
        org_id = line
        username = lines[i - 1] if i - 1 >= 0 else ""
        # El nombre es la línea previa que contiene coma (APELLIDOS, Nombres).
        name = ""
        for j in (i - 2, i - 3):
            if j >= 0 and "," in lines[j] and not _ORG_ID_RE.match(lines[j]):
                name = lines[j]
                break
        if not name and i - 2 >= 0:
            name = lines[i - 2]
        name = name.strip()
        if not name or _ORG_ID_RE.match(name):
            continue
        if org_id in seen:
            continue
        seen.add(org_id)
        students.append({
            "name": name,
            "username": username.strip(),
            "org_id": org_id,
        })
    return students


@app.route("/admin/api/courses", methods=["GET"])
@auth.admin_required
def admin_list_courses():
    return jsonify({"ok": True, "courses": auth.list_courses()})


@app.route("/admin/api/courses/preview", methods=["POST"])
@auth.admin_required
def admin_preview_course():
    data = request.get_json(force=True) or {}
    students = parse_class_list(str(data.get("class_list", "")))
    return jsonify({"ok": True, "students": students, "count": len(students)})


@app.route("/admin/api/courses", methods=["POST"])
@auth.admin_required
def admin_create_course():
    data = request.get_json(force=True) or {}
    name = str(data.get("name", "")).strip()
    if not name:
        return jsonify({"error": "El nombre del curso es obligatorio."}), 400
    students = parse_class_list(str(data.get("class_list", "")))
    if not students:
        return jsonify({"error": "No se detectaron estudiantes en la lista pegada."}), 400
    course_id = auth.create_course(name, session.get("user"), students)
    auth.log_event(session.get("user"), client_ip(), "admin_create_course",
                   detail=f"{name} ({len(students)} estudiantes)")
    return jsonify({"ok": True, "id": course_id, "count": len(students)})


@app.route("/admin/api/courses/<int:course_id>", methods=["GET"])
@auth.admin_required
def admin_get_course(course_id: int):
    course = auth.get_course(course_id)
    if not course:
        return jsonify({"error": "Curso no encontrado."}), 404
    course["students"] = auth.get_course_students(course_id)
    return jsonify({"ok": True, "course": course})


@app.route("/admin/api/courses/<int:course_id>", methods=["DELETE"])
@auth.admin_required
def admin_delete_course(course_id: int):
    if not auth.delete_course(course_id):
        return jsonify({"error": "Curso no encontrado."}), 404
    auth.log_event(session.get("user"), client_ip(), "admin_delete_course", detail=str(course_id))
    return jsonify({"ok": True})


@app.route("/admin/api/courses/<int:course_id>/students", methods=["POST"])
@auth.admin_required
def admin_update_course_students(course_id: int):
    if not auth.get_course(course_id):
        return jsonify({"error": "Curso no encontrado."}), 404
    data = request.get_json(force=True) or {}
    students = parse_class_list(str(data.get("class_list", "")))
    if not students:
        return jsonify({"error": "No se detectaron estudiantes en la lista pegada."}), 400
    auth.set_course_students(course_id, students)
    auth.log_event(session.get("user"), client_ip(), "admin_update_course_students",
                   detail=f"{course_id} ({len(students)} estudiantes)")
    return jsonify({"ok": True, "count": len(students)})


@app.route("/admin/api/users/<username>/course", methods=["POST"])
@auth.admin_required
def admin_assign_course(username: str):
    uname = username.strip().lower()
    if not auth.get_user(uname):
        return jsonify({"error": "Usuario no encontrado."}), 404
    data = request.get_json(force=True) or {}
    course_id = data.get("course_id")
    if course_id in ("", None):
        course_id = None
    else:
        try:
            course_id = int(course_id)
        except (TypeError, ValueError):
            return jsonify({"error": "Curso inválido."}), 400
        if not auth.get_course(course_id):
            return jsonify({"error": "Curso no encontrado."}), 404
    auth.assign_user_course(uname, course_id)
    auth.log_event(session.get("user"), client_ip(), "admin_assign_course",
                   detail=f"{uname}:{course_id}")
    return jsonify({"ok": True})


@app.route("/course/select", methods=["POST"])
@auth.login_required
def select_course():
    """Un admin elige libremente el curso que está calificando."""
    username = session["user"]
    if not auth.get_user(username)["is_admin"]:
        return jsonify({"error": "Solo los administradores pueden elegir su curso."}), 403
    data = request.get_json(silent=True) or {}
    course_id = data.get("course_id")
    if course_id in ("", None):
        course_id = None
    else:
        try:
            course_id = int(course_id)
        except (TypeError, ValueError):
            return jsonify({"error": "Curso inválido."}), 400
        if not auth.get_course(course_id):
            return jsonify({"error": "Curso no encontrado."}), 404
    auth.assign_user_course(username, course_id)
    return jsonify({"ok": True})


# ==============================
# ENTRY POINT
# ==============================

if __name__ == "__main__":
    # Solo abrir el navegador cuando se ejecuta directamente (no en Docker)
    if not os.environ.get("PREGRADER_CONFIG_DIR", "").startswith("/app"):
        threading.Timer(1.2, lambda: webbrowser.open("http://127.0.0.1:5000")).start()
    print("Academic Pregrader en http://127.0.0.1:5000 ...")
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
